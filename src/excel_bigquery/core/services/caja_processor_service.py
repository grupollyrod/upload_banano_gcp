from typing import List
from openpyxl import load_workbook
import logging

from src.excel_bigquery.core.domain.models.caja_model import CajaModel
from src.excel_bigquery.core.domain.models.archivo_model import ArchivoModel
from src.excel_bigquery.core.utils.date_utils import extract_week_and_year_from_trazabilidad
from src.config.settings import settings

logger = logging.getLogger(__name__)


class CajaProcessorService:

    def process_cajas_from_file(self, archivo_model: ArchivoModel, ruta_archivo: str) -> List[CajaModel]:
        """
        Procesa las cajas de un archivo Excel específico

        Args:
            archivo_model: Modelo del archivo que contiene metadata
            ruta_archivo: Ruta completa del archivo Excel

        Returns:
            Lista de modelos CajaModel
        """
        try:
            wb_obj = load_workbook(ruta_archivo)
            sheet_obj = wb_obj.active

            cajas = []

            # Empezar desde la columna B (columna 2) y revisar cada 2 columnas: B, D, F, H, etc.
            column = 2  # Columna B
            max_columns_to_check = 50  # Límite de seguridad

            while column <= max_columns_to_check:
                try:
                    # Verificar si hay nombre de caja en la columna actual, fila 2
                    nombre_caja = sheet_obj.cell(row=2, column=column).value

                    if nombre_caja and str(nombre_caja).strip():  # Si hay datos válidos
                        caja_data = self._extract_caja_data_from_column(sheet_obj, column, archivo_model)
                        if caja_data:
                            cajas.append(caja_data)
                    else:
                        # Si no hay más cajas, salir del bucle
                        break

                    # Avanzar a la siguiente columna par (B -> D -> F -> H, etc.)
                    column += 2

                except Exception as e:
                    logger.warning(f"Error procesando columna {column} del archivo {archivo_model.archivo}: {e}")
                    column += 2
                    continue

            logger.info(f"Procesadas {len(cajas)} cajas del archivo {archivo_model.archivo}")
            return cajas

        except Exception as e:
            logger.error(f"Error procesando cajas del archivo {archivo_model.archivo}: {e}")
            return []

    def _extract_caja_data_from_column(self, sheet_obj, column: int, archivo_model: ArchivoModel) -> CajaModel:
        """
        Extrae los datos de una caja de una columna específica

        Args:
            sheet_obj: Hoja de Excel
            column: Número de columna (B=2, D=4, F=6, etc.)
            archivo_model: Modelo del archivo

        Returns:
            Modelo CajaModel
        """
        # Extraer datos básicos de la caja según las filas especificadas:
        nombre_caja = str(sheet_obj.cell(row=2, column=column).value or "").strip()
        codigo_container = str(sheet_obj.cell(row=3, column=column).value or "").strip()
        codigo_hacienda = self._safe_int(sheet_obj.cell(row=4, column=column).value)
        codigo_trazabilidad = str(sheet_obj.cell(row=5, column=column).value or "").strip()
        nombre_hacienda = str(sheet_obj.cell(row=6, column=column).value or "").strip().upper()  # MAYÚSCULAS

        # Campos que solo pertenecen a NITTSU MATHIAS
        # Para NITTSU, KOBE y HAKATA son 0
        temperatura = 0.0
        dedos_totales = 0
        peso_bruto_kg = 0.0
        cantidad_observaciones = 0
        dedos_afectados_totales = 0

        # Procesar pesos desde la fila 8 hacia abajo según spec_value
        peso_data = self._extract_peso_data_from_column(sheet_obj, column, archivo_model)

        # Extraer semana y año del código de trazabilidad
        week_code, year_code = extract_week_and_year_from_trazabilidad(codigo_trazabilidad)

        # Crear ID único para la caja según el formato especificado:
        # warehouse_anio_archivo_caja_codigocontainer_codigohacienda_codigotrazabilidad_nombrehacienda
        id_caja = f"{archivo_model.warehouse}_{archivo_model.annio}_{archivo_model.archivo}_{nombre_caja.replace(' ', '_')}_{codigo_container}_{codigo_hacienda}_{codigo_trazabilidad}_{nombre_hacienda.replace(' ', '_')}"

        return CajaModel(
            id_caja=id_caja,
            id_archivo=archivo_model.id_archivo,
            nombre_caja=nombre_caja,
            codigo_container=codigo_container,
            codigo_hacienda=codigo_hacienda,
            codigo_trazabilidad=codigo_trazabilidad,
            nombre_hacienda=nombre_hacienda,
            temperatura=round(temperatura, 2),
            dedos_totales=dedos_totales,
            peso_bruto_kg=round(peso_bruto_kg, 2),
            peso_total_kg=round(peso_data['peso_total'], 2),
            cantidad_observaciones=cantidad_observaciones,
            dedos_afectados_totales=dedos_afectados_totales,
            peso_promedio=round(peso_data['peso_promedio'], 2),
            week_code=week_code,
            year_code=year_code,
            spec=archivo_model.spec,
            uw=peso_data['uw'],
            ow=peso_data['ow']
        )

    def _extract_peso_data_from_column(self, sheet_obj, column: int, archivo_model: ArchivoModel) -> dict:
        """
        Extrae los datos de peso de una columna desde la fila 8 hacia abajo

        Args:
            sheet_obj: Hoja de Excel
            column: Número de columna
            archivo_model: Modelo del archivo

        Returns:
            Diccionario con datos de peso, UW y OW
        """
        pesos = []
        uw_count = 0
        ow_count = 0

        # Leer desde la fila 8 hasta fila 8 + spec_value
        fila_inicial = 8
        fila_final = fila_inicial + settings.spec_value  # Por ejemplo: 8 + 30 = 38

        for fila in range(fila_inicial, fila_final):
            try:
                valor_celda = sheet_obj.cell(row=fila, column=column).value

                if valor_celda is not None:
                    peso = float(valor_celda)
                    pesos.append(peso)

                    # Contar UW y OW según los umbrales
                    if peso < settings.uw_threshold:  # < 560
                        uw_count += 1
                    if peso > settings.ow_threshold:  # > 725
                        ow_count += 1

            except (ValueError, TypeError):
                # Si no se puede convertir a float, continuar
                continue

        # Calcular totales y promedio
        peso_total = sum(pesos) if pesos else 0.0
        peso_promedio = peso_total / len(pesos) if pesos else 0.0

        return {
            'peso_total': peso_total,
            'peso_promedio': peso_promedio,
            'uw': uw_count,
            'ow': ow_count,
            'cantidad_pesos': len(pesos)
        }

    def _safe_int(self, value) -> int:
        """Convierte valor a int de forma segura"""
        try:
            if value is None:
                return 0
            return int(float(str(value)))
        except (ValueError, TypeError):
            return 0

    def process_all_cajas_from_files(self, archivos_models: List[ArchivoModel], rutas_archivos: List[str]) -> List[
        CajaModel]:
        """
        Procesa cajas de múltiples archivos

        Args:
            archivos_models: Lista de modelos de archivos
            rutas_archivos: Lista de rutas de archivos correspondientes

        Returns:
            Lista consolidada de todas las cajas
        """
        all_cajas = []

        for archivo_model, ruta_archivo in zip(archivos_models, rutas_archivos):
            cajas = self.process_cajas_from_file(archivo_model, ruta_archivo)
            all_cajas.extend(cajas)

        logger.info(f"Total de cajas procesadas: {len(all_cajas)}")
        return all_cajas